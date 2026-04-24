"""
営業リスト作成アプリ
- ローカルFlaskサーバー（localhost:5000）
- 会社名+URLを入力 → メアド/フォーム/備考を取得 → CSV/Excel DL
- SSEで進捗リアルタイム表示
"""
import csv
import io
import json
import os
import re
import threading
import time
import uuid
from datetime import datetime
from pathlib import Path

from flask import Flask, jsonify, render_template, request, Response, send_file

import scraper

BASE_DIR = Path(__file__).parent
RESULTS_DIR = BASE_DIR / "data" / "results"
RESULTS_DIR.mkdir(parents=True, exist_ok=True)

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 8 * 1024 * 1024  # 8MB

# job管理（メモリ内）
JOBS: dict[str, dict] = {}
JOBS_LOCK = threading.Lock()


def parse_input(text: str, default_category: str = "") -> list[dict]:
    """入力テキスト（1行1社、タブ/カンマ区切り）をパース"""
    companies = []
    lines = [ln.strip() for ln in text.splitlines() if ln.strip() and not ln.strip().startswith("#")]
    for ln in lines:
        # タブ優先、ダメならカンマ
        parts = re.split(r"\t|,", ln)
        parts = [p.strip() for p in parts if p.strip()]
        if len(parts) < 2:
            continue
        name, url = parts[0], parts[1]
        cat = parts[2] if len(parts) > 2 else default_category
        companies.append({"name": name, "url": url, "category": cat})
    return companies


def run_job(job_id: str, companies: list[dict], delay_sec: float):
    """バックグラウンドでスクレイピング実行"""
    def progress(i, total, name, row):
        with JOBS_LOCK:
            JOBS[job_id]["progress"] = {"current": i, "total": total, "name": name}
            JOBS[job_id]["results"].append(row)

    try:
        scraper.process_batch(companies, progress_callback=progress, delay_sec=delay_sec)
        with JOBS_LOCK:
            JOBS[job_id]["status"] = "completed"
            JOBS[job_id]["finished_at"] = datetime.now().isoformat()

        # CSVファイル保存
        csv_path = RESULTS_DIR / f"{job_id}.csv"
        with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, fieldnames=["会社名", "公式サイトURL", "メアド", "問い合わせフォームURL", "業種", "備考"])
            w.writeheader()
            w.writerows(JOBS[job_id]["results"])
    except Exception as e:
        with JOBS_LOCK:
            JOBS[job_id]["status"] = "error"
            JOBS[job_id]["error"] = str(e)


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/manual")
def manual():
    return send_file(BASE_DIR / "manual.html")


@app.route("/api/start", methods=["POST"])
def api_start():
    data = request.get_json(silent=True) or request.form
    text = data.get("text", "").strip() if data else ""
    default_category = (data.get("category") if data else "") or ""
    delay_sec = float(data.get("delay", 2.0)) if data else 2.0

    if not text:
        return jsonify({"error": "入力テキストが空です"}), 400

    companies = parse_input(text, default_category)
    if not companies:
        return jsonify({"error": "有効な行が見つかりません（1行あたり「会社名,URL」形式が必要）"}), 400

    job_id = uuid.uuid4().hex[:12]
    with JOBS_LOCK:
        JOBS[job_id] = {
            "status": "running",
            "started_at": datetime.now().isoformat(),
            "progress": {"current": 0, "total": len(companies), "name": ""},
            "results": [],
            "count": len(companies),
        }

    t = threading.Thread(target=run_job, args=(job_id, companies, delay_sec), daemon=True)
    t.start()

    return jsonify({"job_id": job_id, "count": len(companies)})


@app.route("/api/status/<job_id>")
def api_status(job_id):
    with JOBS_LOCK:
        job = JOBS.get(job_id)
    if not job:
        return jsonify({"error": "job not found"}), 404
    return jsonify({
        "status": job["status"],
        "progress": job["progress"],
        "results": job["results"],
        "count": job["count"],
    })


@app.route("/api/stream/<job_id>")
def api_stream(job_id):
    """SSEストリーム: 進捗リアルタイム配信"""
    def event_stream():
        last_idx = 0
        while True:
            with JOBS_LOCK:
                job = JOBS.get(job_id)
            if not job:
                yield f"event: error\ndata: {json.dumps({'error': 'job not found'})}\n\n"
                return
            # 新しい結果だけ送る
            results = job["results"]
            while last_idx < len(results):
                row = results[last_idx]
                last_idx += 1
                yield f"event: row\ndata: {json.dumps({'index': last_idx, 'total': job['count'], 'row': row})}\n\n"
            if job["status"] == "completed":
                yield f"event: done\ndata: {json.dumps({'count': len(results)})}\n\n"
                return
            if job["status"] == "error":
                yield f"event: error\ndata: {json.dumps({'error': job.get('error', 'unknown')})}\n\n"
                return
            time.sleep(0.5)

    return Response(event_stream(), mimetype="text/event-stream")


@app.route("/api/download/<job_id>.<fmt>")
def api_download(job_id, fmt):
    with JOBS_LOCK:
        job = JOBS.get(job_id)
    if not job:
        return jsonify({"error": "job not found"}), 404
    if job["status"] != "completed":
        return jsonify({"error": "job not finished"}), 400

    fieldnames = ["会社名", "公式サイトURL", "メアド", "問い合わせフォームURL", "業種", "備考"]
    filename = f"営業リスト_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

    if fmt == "csv":
        csv_path = RESULTS_DIR / f"{job_id}.csv"
        return send_file(csv_path, as_attachment=True, download_name=f"{filename}.csv", mimetype="text/csv; charset=utf-8-sig")
    elif fmt == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "営業リスト"
        # ヘッダー
        header_fill = PatternFill(start_color="D97757", end_color="D97757", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        border = Border(left=Side(style="thin", color="DDDDDD"), right=Side(style="thin", color="DDDDDD"),
                        top=Side(style="thin", color="DDDDDD"), bottom=Side(style="thin", color="DDDDDD"))
        for c_idx, h in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=c_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        for r_idx, row in enumerate(job["results"], 2):
            mail = row.get("メアド", "")
            form = row.get("問い合わせフォームURL", "")
            if mail:
                bg = "E8F5E9"
            elif form:
                bg = "FFFDE7"
            else:
                bg = "FFEBEE"
            for c_idx, h in enumerate(fieldnames, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=row.get(h, ""))
                cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
                cell.alignment = Alignment(wrap_text=True, vertical="center")
                cell.border = border
        widths = [28, 40, 55, 50, 20, 30]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        return send_file(bio, as_attachment=True, download_name=f"{filename}.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        return jsonify({"error": "unsupported format"}), 400


if __name__ == "__main__":
    print("=" * 50)
    print("  営業リスト作成アプリ 起動")
    print("  http://localhost:5000 をブラウザで開いてください")
    print("  終了は Ctrl+C")
    print("=" * 50)
    app.run(host="127.0.0.1", port=5000, debug=False, threaded=True)
